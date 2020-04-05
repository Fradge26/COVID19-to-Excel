import json
import os
import sys

import openpyxl
import pandas as pd
import requests
from openpyxl.utils.dataframe import dataframe_to_rows

if getattr(sys, 'frozen', False):
    application_path = os.path.dirname(sys.argv[0])
else:
    application_path = os.path.dirname(os.path.abspath(__file__))


class CovidAPI:

    def __init__(self, key="4b62ec9749mshc6bdc51c6a264ccp1cc1d7jsn2710b1a10b71", excel_path="COVID19_data_book.xlsx"):
        self.url = "https://covid-19-coronavirus-statistics.p.rapidapi.com/v1/stats"
        self.headers = {
            'x-rapidapi-host': "covid-19-coronavirus-statistics.p.rapidapi.com",
            'x-rapidapi-key': key
        }
        self.excel_path = os.path.join(application_path, excel_path)

    def get_json_data(self):
        response = requests.request("GET", self.url, headers=self.headers)
        json_data = json.loads(response.text)
        return json_data

    @staticmethod
    def json_to_df(json_data):
        data_df = pd.json_normalize(json_data["data"]["covid19Stats"])
        data_df = data_df[["country", "province", "confirmed", "recovered", "deaths", "lastUpdate"]]
        return data_df

    @staticmethod
    def json_to_timestamp(json_data):
        return json_data["data"]["lastChecked"]

    def open_xls(self):
        if os.path.exists(self.excel_path):
            wb = openpyxl.load_workbook(self.excel_path)
        else:
            wb = openpyxl.Workbook()
            wb.remove(wb["Sheet"])
        return wb

    def update_workbook(self):
        json_data = self.get_json_data()

        province_data = self.json_to_df(json_data)
        country_data = province_data.groupby("country").sum()
        country_data = country_data.sort_values(by="confirmed", ascending=False)
        country_data = country_data.reset_index()
        data = {"province": province_data,
                "country": country_data}

        timestamp = self.json_to_timestamp(json_data)
        date_string = timestamp[:10]
        wb = self.open_xls()
        for prefix in ["country", "province"]:
            sheet_name = f"{prefix} {date_string}"
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(f"{prefix} {date_string}")
            self.df_to_sheet(data[prefix], ws, timestamp)
        wb.save(filename=self.excel_path)
        print(f"success: COVID-19 data written to {self.excel_path}")

    @staticmethod
    def df_to_sheet(df, ws, timestamp):
        if timestamp:
            ws["A1"] = f"data updated on: {timestamp}"
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)


if __name__ == "__main__":
    covid = CovidAPI()
    covid.update_workbook()
